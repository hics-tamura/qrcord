from flask import Flask, render_template, request
from PIL import Image
import openpyxl as px
import qrcode
import os
import shutil
import re
import datetime

'''
=================================================================
                          共通処理
=================================================================
'''

app = Flask(__name__)
if __name__ == '__main__':
    app.run()
    # app.run(host="0.0.0.0", port="8080")

# Top画面表示処理の実装メソッド
@app.route('/')
def index():
    return render_template('top.html')

# エラー画面表示処理の実装メソッド
@app.route('/error.html')
def error():
    return render_template('error.html')


'''
=================================================================
                          QRコード作成処理
=================================================================
'''

# QRコード作成エリア選択画面表示処理の実装メソッド
@app.route('/create/createAreaSelect.html')
def createAreaSelect():
    return render_template('/create/createAreaSelect.html')

# QRコード作成処理の実装メソッド
# 作成エリア選択画面で選択したエリアに紐づくQRコードを作成する。
@app.route("/create/createComplete.html", methods=["POST"])
def create():
    count = 0
    qr_list = []
    get_value = request.form.getlist("plase")
    plase = get_value[0]
    qr_dir = 'image/qrcode/' + plase
    book = px.load_workbook('excel/管理台帳＆一覧表 - 印刷用.xlsx')
    now = book.active
    # デーらが入っているマックス行をとる
    row = book['ハードウェア台帳'].max_row
    # データが入っているマックス列をとる
    col = book['ハードウェア台帳'].max_column
    # ここでフォルダが無ければ作る
    # フォルダがあれば削除してから作り直すをしている

    if not os.path.exists(qr_dir):
        os.makedirs(qr_dir)
    else:
        shutil.rmtree(qr_dir)
        os.makedirs(qr_dir)
    # 選択した場所をもとに台帳から取り出す
    for i in range(row):
        sheet = book['ハードウェア台帳']
        cell = "H" + str(i + 1)
        if sheet[cell].value == plase:
            manegecell = "C" + str(i + 1)
            qr = qrcode.QRCode(box_size=3)
            buf = sheet[manegecell].value
            buf = buf.replace('書架カメラ', 'shokacamera')
            buf = buf.replace('セット', 'set')
            qr.add_data(buf)
            qr.make()
            img_qr = qr.make_image()
            qr_list.append(sheet[manegecell].value)
            manegeNo = sheet[manegecell].value
            manegeNo = manegeNo.replace('/', '／')
            img_qr.save(qr_dir + '/' + manegeNo + '.png')
            count += 1

    return render_template('/create/createComplete.html', plase=plase,
                            count=count, qr_list=qr_list)


'''
=================================================================
                          管理番号照合処理
=================================================================
'''

# 読取エリア選択画面表示処理の実装メソッド
@app.route('/match/readAreaSelect.html')
def readAreaSelect():
    return render_template('/match/readAreaSelect.html')

# 管理番号抽出処理の実装メソッド
# 端末から読取エリア選択画面で選択したエリアに紐づく管理番号を抽出する。
@app.route("/match/controlNumberReadingList.html", methods=["POST"])
def read():
    count = 0
    qr_list = []
    counts = []
    get_value = request.form.getlist("plase")
    plase = get_value[0]
    book = px.load_workbook('excel/管理台帳＆一覧表 - 印刷用.xlsx')
    now = book.active
    # マックス行をとる
    row = book['ハードウェア台帳'].max_row
    for i in range(row):
        sheet = book['ハードウェア台帳']
        cell = "H" + str(i + 1)
        if sheet[cell].value == plase:
            manegecell = "C" + str(i + 1)
            qr_list.append(sheet[manegecell].value)
            count += 1
            counts.append(count)
    plus = counts[len(counts) - 1]
    for i in range(150):
        counts.append(i + plus)
    return render_template('/match/controlNumberReadingList.html',
                            plase=plase, qr_list=qr_list, counts=counts)

# 管理番号突合処理の実装メソッド
# 管理台帳から取得した管理番号とQRコードから読み取った管理番号の突き合わせを行う。
@app.route("/match/contrloNumberMatchResult.html", methods=["POST"])
def match():
    read_qr_list = request.form.getlist("matching")
    get_value = request.form.getlist("plase")
    plase = get_value[0]
    mg_list = []
    book = px.load_workbook('excel/管理台帳＆一覧表 - 印刷用.xlsx')
    now = book.active
    history_book = px.load_workbook('excel/履歴.xlsx')
    now2 = history_book.active
    # マックス行をとる
    row = book['ハードウェア台帳'].max_row
    dt_now = datetime.datetime.now()
    c_detail_list = []
    c_machine_list = []
    for i in range(len(read_qr_list)):
        read_qr_list[i] = read_qr_list[i].replace('shokacamera', '書架カメラ')
        read_qr_list[i] = read_qr_list[i].replace('set', 'セット')
    for i in range(row):
        sheet = book['ハードウェア台帳']
        cell = "H" + str(i + 1)
        if sheet[cell].value == plase:
            manegecell = "C" + str(i + 1)
            mg_list.append(sheet[manegecell].value)
            cell1 = 'E' + str(i + 1)
            c_machine_list.append(sheet[cell1].value)
            cell2 = 'I' + str(i + 1)
            c_detail_list.append(sheet[cell2].value)
    c_list = []
    d_list = []
    d_place_list = []

    count = 0
    for i in range(len(mg_list)):
        c_list.append("")
    for chack in read_qr_list:
        for mg_index, lastflg in lastone(range(len(mg_list))):
            if chack == mg_list[mg_index]:
                c_list[mg_index] = mg_list[mg_index]
                break
            if lastflg:
                d_list.append(chack)
                break
    for i in range(row):
        for d_value in d_list:
            sheet = book['ハードウェア台帳']
            cell = "C" + str(i + 1)
            if sheet[cell].value == d_value:
                cell = "H" + str(i + 1)
                d_place_list.append(sheet[cell].value)

    history_newsheet = history_book.create_sheet(
                            str(plase + '_' + str(datetime.date.today())))
    for i in range(len(read_qr_list)):
        history_cell = 'A' + str(i + 1)
        history_newsheet[history_cell] = read_qr_list[i]
    history_book.save('excel/履歴.xlsx')
    d_list = list(set(d_list))
    d_list.sort()
    d_list.remove("")
    d_place_list = []
    d_machine_list = []
    d_detail_list = []
    print(d_list)
    for i in range(row):
        sheet = book['ハードウェア台帳']
        cell = 'C' + str(i + 1)
        for j in range(len(d_list)):
            if sheet[cell].value == d_list[j]:
                cell1 = 'H' + str(i + 1)
                d_place_list.append(sheet[cell1].value)
                cell2 = 'E' + str(i + 1)
                d_detail_list.append(sheet[cell2].value)
                cell3 = 'I' + str(i + 1)
                d_machine_list.append(sheet[cell3].value)

    return render_template('/match/contrloNumberMatchResult.html',
                            qr_list=c_list, mg_list=mg_list, plase=plase,
                            d_list=d_list, c_detail_list=c_detail_list,
                            c_machine_list=c_machine_list,
                            d_place_list=d_place_list,
                            d_detail_list=d_detail_list,
                            d_machine_list=d_machine_list,
                            mg_list_lng=len(mg_list))

# 与えられたイテレータブルオブジェクトの
# 最後の一つの要素の時にTrue、それ以外の時にFalseを返す
def lastone(iterable):
    # イテレータを取得して最初の値を取得する
    it = iter(iterable)
    last = next(it)
    # 2番目の値から開始して反復子を使い果たすまで実行
    for val in it:
        # 一つ前の値を返す
        yield last, False
        last = val  # 値の更新
    # 最後の一つ
    yield last, True


'''
=================================================================
                          管理台帳更新処理
=================================================================
'''

#
#
@app.route("/update.html", methods=["POST"])
def update():
    d_place_list = request.form.getlist("d_place")
    d_code_list = request.form.getlist("d_code")
    read_place = request.form.getlist("place")
    place = read_place[0]
    book = px.load_workbook('excel/管理台帳＆一覧表 - 印刷用.xlsx')
    now = book.active
    # マックス行をとる
    row = book['ハードウェア台帳'].max_row
    sheet = book['ハードウェア台帳']
    for i in range(row):
        cell = "C" + str(i + 1)
        for j in range(len(d_code_list)):
            if d_code_list[j] == sheet[cell].value:
                changecell = "H" + str(i + 1)
                sheet[changecell] = place
                sheet[changecell].font = px.styles.fonts.Font(color='FF0000')
    book.save('excel/管理台帳＆一覧表_更新.xlsx')
    return render_template('update.html', d_place_list=d_place_list,
                            d_code_list=d_code_list, place=place)
